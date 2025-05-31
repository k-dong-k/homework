import os
# Disable MPS (Apple Silicon GPU) to enforce CPU usage and avoid device mismatch
os.environ["PYTORCH_ENABLE_MPS_FALLBACK"] = "0"

import torch
# Disable MPS backend if available
def disable_mps():
    if hasattr(torch.backends, "mps"):
        torch.backends.mps.is_available = lambda: False
        torch.backends.mps.is_built = lambda: False

disable_mps()

# Patch for Streamlit-Torch conflict: avoid introspection error
if hasattr(torch, "classes") and hasattr(torch.classes, "__path__"):
    torch.classes.__path__ = []

import streamlit as st
import zipfile
import easyocr
import numpy as np
import cv2
from PIL import Image, ExifTags
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import uuid
import hashlib

# Streamlit 페이지 설정
st.set_page_config(layout='wide')

@st.cache_resource

def load_model():
    """모델 로드: CPU 환경에서 YOLOv5 차량 탐지, 사용자 학습된 번호판 탐지, EasyOCR 리더 캐싱"""
    device = torch.device("cpu")
    
    # ⚠️ GitHub 대신 로컬에서 로드
    car_m = torch.hub.load(
        './yolov5', 'yolov5s', source='local'
    ).to(device)

    lp_m = torch.hub.load(
        './yolov5', 'custom', path='lp_det.pt', source='local'
    ).to(device)

    reader = easyocr.Reader(
        ['en'], detect_network='craft', recog_network='best_acc',
        user_network_directory='lp_models/user_network',
        model_storage_directory='lp_models/models',
        gpu=False
    )
    car_m.classes = [2, 3, 5, 7]
    return car_m, lp_m, reader

def detect(car_m, lp_m, reader, path):
    """차량 탐지 → 번호판 영역 검출 → OCR, 실패 시 전체 이미지 OCR로 재시도"""
    # 1) 원본 이미지 로드
    im_pil = Image.open(path).convert("RGB")
    img = np.array(im_pil)

    # 2) 차량 탐지
    results = car_m(im_pil)
    locs = results.xyxy[0]
    result_text = []

    # 3) 차량 영역 내 번호판 검출→OCR
    for item in locs:
        x1, y1, x2, y2 = [int(t.cpu().detach().numpy()) for t in item[:4]]
        car_crop = img[y1:y2, x1:x2, :].copy()
        lp_results = lp_m(Image.fromarray(car_crop))
        for lp in lp_results.xyxy[0]:
            lx1, ly1, lx2, ly2 = [int(t.cpu().detach().numpy()) for t in lp[:4]]
            plate_crop = car_crop[ly1:ly2, lx1:lx2]
            resized = cv2.resize(plate_crop, (224, 128))
            gray = cv2.cvtColor(resized, cv2.COLOR_BGR2GRAY)
            ocr = reader.recognize(gray)
            if ocr:
                result_text.append(ocr[0][1])

    # 4) 차량 영역에서 인식 실패 시 → 전체 이미지 OCR 재시도
    if not result_text:
        full_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        ocr_full = reader.recognize(full_gray)
        if ocr_full:
            result_text = [res[1] for res in ocr_full]
        else:
            result_text = ["인식 실패"]

    return path, result_text

def get_image_date(path):
    """EXIF 또는 파일 수정 시간으로부터 촬영일 추출"""
    try:
        img = Image.open(path)
        exif = img._getexif()
        if exif:
            for tag, val in exif.items():
                name = ExifTags.TAGS.get(tag)
                if name in ('DateTimeOriginal', 'DateTime'):
                    return datetime.strptime(val, '%Y:%m:%d %H:%M:%S')
    except Exception:
        pass
    return datetime.fromtimestamp(os.path.getmtime(path))

def save_uploaded_file(directory, uploaded_file):
    os.makedirs(directory, exist_ok=True)
    path = os.path.join(directory, uploaded_file.name)
    with open(path, 'wb') as f:
        f.write(uploaded_file.getbuffer())
    return path

def extract_zip(zip_path):
    """ZIP 압축 해제 (고유 폴더)"""
    extract_to = f"extracted/{uuid.uuid4().hex}"
    os.makedirs(extract_to, exist_ok=True)
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)
    return extract_to

def get_file_hash(bytes_data):
    return hashlib.md5(bytes_data).hexdigest()

def save_to_excel(infos, filename):
    """엑셀 저장"""
    os.makedirs('excel_outputs', exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = f"excel_outputs/{filename}_{timestamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "차량 인식 결과"
    ws.append(["촬영일", "파일명", "차량 번호", "이미지 미리보기"])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40

    for idx, info in enumerate(infos, start=2):
        ws.cell(idx, 1, info['capture_time'])
        ws.cell(idx, 2, info['name'])
        ws.cell(idx, 3, info['plate'])
        try:
            img = ExcelImage(info['path'])
            img.width = 150
            img.height = int(150 * img.height / img.width)
            ws.add_image(img, f"D{idx}")
            ws.row_dimensions[idx].height = img.height * 0.75
        except Exception as e:
            print("이미지 삽입 실패", e)

    wb.save(out_path)
    st.success(f"엑셀 저장 완료: {out_path}")

def main():
    car_m, lp_m, reader = load_model()

    if 'uploaded_hashes' not in st.session_state:
        st.session_state['uploaded_hashes'] = set()

    if 'file_info' not in st.session_state:
        st.session_state['file_info'] = []

    st.title("🚗 차량 번호판 자동 인식 시스템")
    menu = ['About', '파일 업로드', '결과 확인 및 수정']
    choice = st.sidebar.selectbox('메뉴', menu)

    if choice == 'About':
        st.markdown("### 개요")
        st.write("차량 이미지 혹은 ZIP 파일 업로드 시 YOLO+OCR로 번호판을 자동 인식합니다.")

    elif choice == '파일 업로드':
        st.markdown("### 이미지 / ZIP 업로드")
        uploaded = st.file_uploader(
            "(PNG, JPG, JPEG, ZIP 지원)",
            type=['png','jpg','jpeg','zip'],
            accept_multiple_files=True
        )
        if uploaded:
            infos = []
            for f in uploaded:
                ext = f.name.split('.')[-1].lower()
                file_hash = get_file_hash(f.getvalue())

                # 중복된 파일이 있으면 기존 정보를 덮어쓰기
                existing_info = None
                for info in st.session_state['file_info']:
                    if get_file_hash(open(info['path'], 'rb').read()) == file_hash:
                        existing_info = info
                        break

                if existing_info:
                    st.warning(f"중복된 파일이 있습니다: {f.name}. 기존 정보를 덮어씁니다.")
                    st.session_state['file_info'].remove(existing_info)

                if ext == 'zip':
                    zp = save_uploaded_file('uploads', f)
                    fld = extract_zip(zp)
                    for root, _, files in os.walk(fld):
                        for fn in files:
                            if fn.lower().endswith(('png','jpg','jpeg')):
                                full_path = os.path.join(root, fn)

                                # detect() 반환값 unpack
                                result_path, plates = detect(car_m, lp_m, reader, full_path)

                                # plates 리스트 flatten
                                flat = []
                                for p in plates:
                                    if isinstance(p, (list, tuple)):
                                        flat.extend(str(x) for x in p)
                                    else:
                                        flat.append(str(p))
                                plate_str = ", ".join(flat)

                                capture_dt = get_image_date(result_path)
                                infos.append({
                                    'capture_time': capture_dt.strftime('%Y-%m-%d %H:%M:%S'),
                                    'name': fn,
                                    'plate': plate_str,
                                    'path': result_path
                                })

                else:
                    saved_path = save_uploaded_file('uploads', f)
                    result_path, plates = detect(car_m, lp_m, reader, saved_path)

                    flat = []
                    for p in plates:
                        if isinstance(p, (list, tuple)):
                            flat.extend(str(x) for x in p)
                        else:
                            flat.append(str(p))
                    plate_str = ", ".join(flat)

                    capture_dt = get_image_date(result_path)
                    infos.append({
                        'capture_time': capture_dt.strftime('%Y-%m-%d %H:%M:%S'),
                        'name': f.name,
                        'plate': plate_str,
                        'path': result_path
                    })

            # 새로운 파일 정보 추가
            st.session_state['file_info'] += infos

            # 촬영일 기준 정렬
            st.session_state['file_info'].sort(key=lambda x: datetime.strptime(x['capture_time'], '%Y-%m-%d %H:%M:%S'))
            st.success("✅ 업로드 완료!")

    else:  # 결과 확인 및 수정
        if not st.session_state['file_info']:
            st.info("업로드 후 확인하세요.")
            return

        file_info = st.session_state['file_info']
        plates = sorted({info['plate'] for info in file_info})
        selected = st.radio("차량 번호 선택", plates)

        for info in file_info:
            if info['plate'] == selected:
                cols = st.columns([2,1,1,1])
                with cols[0]:
                    st.image(info['path'], use_container_width=True)
                with cols[1]:
                    st.write("기존:", info['plate'])
                with cols[2]:
                    new = st.text_input("수정", key=info['path'])
                with cols[3]:
                    if st.button("저장", key=info['path'] + "_btn"):
                        info['plate'] = new
                        st.session_state['file_info'] = file_info  # 수정된 데이터 업데이트
                        st.experimental_rerun()

        st.markdown("---")
        fn = st.text_input("엑셀 파일명", "vehicles")
        if st.button("엑셀 저장"):
            save_to_excel(file_info, fn)
        
if __name__ == '__main__':
    main() 